from operator import index
from turtle import color
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from collections import Counter


def highlighting():

    # open file
    excel = "excel_docs\in_progress.xlsx"
    # openpyxl way of opening a workbook
    wb = load_workbook(excel)
    ws = wb.active
    # highlight fave books yellow
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    # highlight fave authors pink
    pink_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    sheet1 = wb["books"]
    sheet2 = wb["favorites"]

    # sheet2 fave books list
    fave_books = []
    for column in sheet2["A"]:
        fave_books.append(column.value)

    # sheet2 authors list
    fave_authors = []
    for column in sheet2["B"]:
        fave_authors.append(column.value)

    # highlighting all the faves_books:
    for rows in sheet1.iter_rows(min_row=1, max_row=15, min_col=0, max_col=3):
        for cell in rows:
            # if statement works now!
            if cell.value in fave_books:
                cell.fill = yellow_fill
                wb.save("my_test.xlsx")

    # highlighting fave authors
    for rows in sheet1.iter_rows(min_row=1, max_row=15, min_col=0, max_col=3):
        for cell in rows:
            # if statement works now!
            if cell.value in fave_authors:
                cell.fill = pink_fill
                wb.save("my_test.xlsx")

    # create a new sheet for the key
    wb.create_sheet("what_does_it_mean")
    sheet3 = wb["what_does_it_mean"]
    # yellow are the books, pink are the authors
    # intro
    intro = sheet3["A1"]
    intro.font = Font(name="Helvetica", size=14, bold=True)
    intro.value = "KEY:"

    # yellow books
    yellow = sheet3["A2"]
    yellow.value = "favorite books"
    yellow.fill = yellow_fill

    # pink books
    pink = sheet3["B2"]
    pink.value = "favorite authors"
    pink.fill = pink_fill

    # save wb
    wb.save("my_test.xlsx")

    # create a new sheet instances of ISBNs
    wb.create_sheet("count")
    sheet4 = wb["count"]

    # turn ISBN list in sheet1 col C to a list
    isbn = []
    for column in sheet1["C"]:
        isbn.append(column.value)
    print(isbn)

    # count number of occurances per ISBN instance
    counted = Counter(isbn)

    wb.create_sheet("teeeest")
    sheet5 = wb["teeeest"]

    # letters = ["a", "b", "c"]
    # num = [1, 2, 3]

    # for x in letters:
    #     print(x)
    # for y in num:
    #     print(y)

    # print all key value pairs (ISBNs with counts)
    # for key, value in counted.items():
    # print(key)
    # k_list.append(key)
    # print(value)
    # v_list.append(key)

    # 14
    # print(len(counted))
    for it in counted:
        for key, value in counted.items():
            keys = sheet5[f"A{it}"]
            keys.value = key
            it += 1
            wb.save("my_test.xlsx")
    print(value)

    # create a new sheet for a test
    # wb.create_sheet("teeeest")
    # sheet5 = wb["teeeest"]

    # letters = ["a", "b", "c"]
    # num = [1, 2, 3]

    # for x in letters:
    #     print(x)
    # for y in num:
    #     print(y)

    # # test = sheet5[""]

    # for it in range(1, 4):
    #     lets = sheet5[f"A{it}"]
    #     lets.value = key
    #     it += 1
    #     wb.save("my_test.xlsx")

    # # create a new sheet for the key
    # wb.create_sheet("what_does_it_mean")
    # sheet3 = wb["what_does_it_mean"]
    # # yellow are the books, pink are the authors
    # # intro
    # intro = sheet3["A1"]
    # intro.font = Font(name="Helvetica", size=14, bold=True)
    # intro.value = "KEY:"


if __name__ == "__main__":
    highlighting()
