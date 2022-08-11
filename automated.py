# use openpyxl, xlsxwriter, pandas
# open excel doc (pd/xlsxwriter), edit (openpyxl)
# bold first line
# add '5' to the first line of the ratings
# add title of my 'book'
# highlight fave books

from openpyxl import load_workbook
import pandas as pd

# open file
to_do_excel = ""
to_do_df = pd.read_excel(to_do_excel)
print(to_do_df)

# fill in first book as 5


writer = pd.ExcelWriter(to_do_excel, engine="openpyxl")
book = load_workbook(fn)
writer.book = load_workbook(to_do_excel)
